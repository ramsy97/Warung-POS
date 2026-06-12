import io
import datetime
from typing import List, Optional
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors

from .database import engine, Base, get_db
from .config import settings
from . import models, schemas, auth, utils, seed

# Initialize FastAPI App
app = FastAPI(title=settings.PROJECT_NAME, version="1.0.0")

# Setup CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)

# Run Seed on Startup
@app.on_event("startup")
def startup_event():
    seed.seed_db()

# ==========================================
# 1. AUTHENTICATION MODULE
# ==========================================

@app.post("/api/auth/login", response_model=schemas.Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == form_data.username).first()
    if not user or not auth.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Username atau password salah",
            headers={"WWW-Authenticate": "Bearer"},
        )
    if not user.is_active:
        raise HTTPException(status_code=400, detail="Akun dinonaktifkan")
        
    access_token = auth.create_access_token(data={"sub": user.username})
    
    # Audit log
    utils.create_audit_log(db, user.id, "Login", "Auth", f"User {user.username} berhasil login.")
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "role": user.role,
        "username": user.username,
        "full_name": user.full_name
    }

# Fallback login using JSON body (useful for web clients)
@app.post("/api/auth/login-json", response_model=schemas.Token)
def login_json(credentials: schemas.UserLogin, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == credentials.username).first()
    if not user or not auth.verify_password(credentials.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Username atau password salah",
            headers={"WWW-Authenticate": "Bearer"},
        )
    if not user.is_active:
        raise HTTPException(status_code=400, detail="Akun dinonaktifkan")
        
    access_token = auth.create_access_token(data={"sub": user.username})
    
    # Audit log
    utils.create_audit_log(db, user.id, "Login", "Auth", f"User {user.username} berhasil login via JSON.")
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "role": user.role,
        "username": user.username,
        "full_name": user.full_name
    }

@app.get("/api/auth/me", response_model=schemas.UserResponse)
def get_me(current_user: models.User = Depends(auth.get_current_user)):
    return current_user


# ==========================================
# 2. USER MANAGEMENT MODULE (Super Admin only)
# ==========================================

@app.get("/api/users", response_model=List[schemas.UserResponse])
def get_users(
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(auth.RoleChecker(["super_admin"]))
):
    return db.query(models.User).all()

@app.post("/api/users", response_model=schemas.UserResponse)
def create_user(
    user_in: schemas.UserCreate, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(auth.RoleChecker(["super_admin"]))
):
    exists = db.query(models.User).filter(models.User.username == user_in.username).first()
    if exists:
        raise HTTPException(status_code=400, detail="Username sudah terdaftar")
    
    hashed_pwd = auth.get_password_hash(user_in.password)
    user = models.User(
        username=user_in.username,
        hashed_password=hashed_pwd,
        full_name=user_in.full_name,
        email=user_in.email,
        role=user_in.role,
        is_active=user_in.is_active
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    
    utils.create_audit_log(db, current_user.id, "Create User", "Users", f"Membuat user baru: {user.username} ({user.role})")
    return user

@app.put("/api/users/{user_id}", response_model=schemas.UserResponse)
def update_user(
    user_id: int, 
    user_in: schemas.UserUpdate, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(auth.RoleChecker(["super_admin"]))
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
        
    if user_in.full_name is not None:
        user.full_name = user_in.full_name
    if user_in.email is not None:
        user.email = user_in.email
    if user_in.role is not None:
        user.role = user_in.role
    if user_in.is_active is not None:
        user.is_active = user_in.is_active
    if user_in.password is not None and user_in.password != "":
        user.hashed_password = auth.get_password_hash(user_in.password)
        
    db.commit()
    db.refresh(user)
    
    utils.create_audit_log(db, current_user.id, "Update User", "Users", f"Memperbarui user: {user.username}")
    return user

@app.delete("/api/users/{user_id}")
def delete_user(
    user_id: int, 
    db: Session = Depends(get_db), 
    current_user: models.User = Depends(auth.RoleChecker(["super_admin"]))
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Tidak dapat menghapus diri sendiri")
        
    db.delete(user)
    db.commit()
    
    utils.create_audit_log(db, current_user.id, "Delete User", "Users", f"Menghapus user: {user.username}")
    return {"message": "User berhasil dihapus"}


# ==========================================
# 3. CATEGORY MODULE (Admin, Owner, Gudang)
# ==========================================

@app.get("/api/categories", response_model=List[schemas.CategoryResponse])
def get_categories(db: Session = Depends(get_db)):
    return db.query(models.Category).all()

@app.post("/api/categories", response_model=schemas.CategoryResponse)
def create_category(
    category_in: schemas.CategoryCreate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "staff_gudang"]))
):
    exists = db.query(models.Category).filter(models.Category.name == category_in.name).first()
    if exists:
        raise HTTPException(status_code=400, detail="Nama kategori sudah ada")
        
    category = models.Category(name=category_in.name, description=category_in.description)
    db.add(category)
    db.commit()
    db.refresh(category)
    
    utils.create_audit_log(db, current_user.id, "Create Category", "Categories", f"Membuat kategori: {category.name}")
    return category

@app.put("/api/categories/{category_id}", response_model=schemas.CategoryResponse)
def update_category(
    category_id: int, 
    category_in: schemas.CategoryCreate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "staff_gudang"]))
):
    category = db.query(models.Category).filter(models.Category.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Kategori tidak ditemukan")
        
    category.name = category_in.name
    category.description = category_in.description
    db.commit()
    db.refresh(category)
    
    utils.create_audit_log(db, current_user.id, "Update Category", "Categories", f"Memperbarui kategori: {category.name}")
    return category

@app.delete("/api/categories/{category_id}")
def delete_category(
    category_id: int, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "staff_gudang"]))
):
    category = db.query(models.Category).filter(models.Category.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Kategori tidak ditemukan")
        
    # Check if category has products
    has_products = db.query(models.Product).filter(models.Product.category_id == category_id).first()
    if has_products:
        raise HTTPException(status_code=400, detail="Kategori tidak bisa dihapus karena masih digunakan oleh produk")
        
    db.delete(category)
    db.commit()
    
    utils.create_audit_log(db, current_user.id, "Delete Category", "Categories", f"Menghapus kategori: {category.name}")
    return {"message": "Kategori berhasil dihapus"}


# ==========================================
# 4. SUPPLIER MODULE (Admin, Owner, Gudang)
# ==========================================

@app.get("/api/suppliers", response_model=List[schemas.SupplierResponse])
def get_suppliers(db: Session = Depends(get_db)):
    return db.query(models.Supplier).all()

@app.post("/api/suppliers", response_model=schemas.SupplierResponse)
def create_supplier(
    supplier_in: schemas.SupplierCreate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "staff_gudang"]))
):
    supplier = models.Supplier(**supplier_in.dict())
    db.add(supplier)
    db.commit()
    db.refresh(supplier)
    
    utils.create_audit_log(db, current_user.id, "Create Supplier", "Suppliers", f"Membuat supplier: {supplier.name}")
    return supplier

@app.put("/api/suppliers/{supplier_id}", response_model=schemas.SupplierResponse)
def update_supplier(
    supplier_id: int, 
    supplier_in: schemas.SupplierCreate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "staff_gudang"]))
):
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier tidak ditemukan")
        
    for k, v in supplier_in.dict().items():
        setattr(supplier, k, v)
        
    db.commit()
    db.refresh(supplier)
    
    utils.create_audit_log(db, current_user.id, "Update Supplier", "Suppliers", f"Memperbarui supplier: {supplier.name}")
    return supplier

@app.delete("/api/suppliers/{supplier_id}")
def delete_supplier(
    supplier_id: int, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "staff_gudang"]))
):
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier tidak ditemukan")
        
    db.delete(supplier)
    db.commit()
    
    utils.create_audit_log(db, current_user.id, "Delete Supplier", "Suppliers", f"Menghapus supplier: {supplier.name}")
    return {"message": "Supplier berhasil dihapus"}

@app.get("/api/suppliers/{supplier_id}/purchases", response_model=List[schemas.PurchaseResponse])
def get_supplier_purchase_history(supplier_id: int, db: Session = Depends(get_db)):
    return db.query(models.Purchase).filter(models.Purchase.supplier_id == supplier_id).order_by(models.Purchase.purchase_date.desc()).all()


# ==========================================
# 5. CUSTOMER MODULE (Admin, Owner, Kasir)
# ==========================================

@app.get("/api/customers", response_model=List[schemas.CustomerResponse])
def get_customers(db: Session = Depends(get_db)):
    return db.query(models.Customer).all()

@app.post("/api/customers", response_model=schemas.CustomerResponse)
def create_customer(
    customer_in: schemas.CustomerCreate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "kasir"]))
):
    customer = models.Customer(**customer_in.dict())
    db.add(customer)
    db.commit()
    db.refresh(customer)
    
    utils.create_audit_log(db, current_user.id, "Create Customer", "Customers", f"Membuat pelanggan: {customer.name}")
    return customer

@app.put("/api/customers/{customer_id}", response_model=schemas.CustomerResponse)
def update_customer(
    customer_id: int, 
    customer_in: schemas.CustomerCreate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "kasir"]))
):
    customer = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Pelanggan tidak ditemukan")
        
    for k, v in customer_in.dict().items():
        setattr(customer, k, v)
        
    db.commit()
    db.refresh(customer)
    
    utils.create_audit_log(db, current_user.id, "Update Customer", "Customers", f"Memperbarui pelanggan: {customer.name}")
    return customer

@app.delete("/api/customers/{customer_id}")
def delete_customer(
    customer_id: int, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "kasir"]))
):
    customer = db.query(models.Customer).filter(models.Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Pelanggan tidak ditemukan")
        
    db.delete(customer)
    db.commit()
    
    utils.create_audit_log(db, current_user.id, "Delete Customer", "Customers", f"Menghapus pelanggan: {customer.name}")
    return {"message": "Pelanggan berhasil dihapus"}

@app.get("/api/customers/{customer_id}/sales", response_model=List[schemas.SaleResponse])
def get_customer_sales_history(customer_id: int, db: Session = Depends(get_db)):
    return db.query(models.Sale).filter(models.Sale.customer_id == customer_id).order_by(models.Sale.sale_date.desc()).all()


# ==========================================
# 6. PRODUCTS MODULE (Admin, Owner, Gudang)
# ==========================================

@app.get("/api/products", response_model=List[schemas.ProductResponse])
def get_products(
    db: Session = Depends(get_db),
    search: Optional[str] = None,
    category_id: Optional[int] = None,
    low_stock: Optional[bool] = False
):
    query = db.query(models.Product)
    if search:
        query = query.filter(
            or_(
                models.Product.name.like(f"%{search}%"),
                models.Product.sku.like(f"%{search}%"),
                models.Product.barcode.like(f"%{search}%")
            )
        )
    if category_id:
        query = query.filter(models.Product.category_id == category_id)
    if low_stock:
        query = query.filter(models.Product.stock <= models.Product.min_stock)
        
    return query.all()

@app.get("/api/products/{product_id}", response_model=schemas.ProductResponse)
def get_product(product_id: int, db: Session = Depends(get_db)):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produk tidak ditemukan")
    return product

@app.post("/api/products", response_model=schemas.ProductResponse)
def create_product(
    product_in: schemas.ProductCreate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "staff_gudang"]))
):
    # Verify SKU uniqueness
    exists_sku = db.query(models.Product).filter(models.Product.sku == product_in.sku).first()
    if exists_sku:
        raise HTTPException(status_code=400, detail="SKU sudah digunakan")
        
    if product_in.barcode:
        exists_barcode = db.query(models.Product).filter(models.Product.barcode == product_in.barcode).first()
        if exists_barcode:
            raise HTTPException(status_code=400, detail="Barcode sudah digunakan")
            
    product = models.Product(**product_in.dict())
    db.add(product)
    db.commit()
    db.refresh(product)
    
    # Stock Movement for initial stock
    if product.stock > 0:
        mov = models.StockMovement(
            product_id=product.id,
            type="Masuk",
            qty=product.stock,
            reference="Create Product",
            notes="Saldo awal produk baru",
            created_by=current_user.id
        )
        db.add(mov)
        db.commit()
        
    utils.check_low_stock(db, product)
    utils.create_audit_log(db, current_user.id, "Create Product", "Products", f"Membuat produk: {product.name} (SKU: {product.sku})")
    
    return product

@app.put("/api/products/{product_id}", response_model=schemas.ProductResponse)
def update_product(
    product_id: int, 
    product_in: schemas.ProductUpdate, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "staff_gudang"]))
):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produk tidak ditemukan")
        
    # Check uniqueness of SKU and Barcode if they are updated
    if product_in.sku and product_in.sku != product.sku:
        exists_sku = db.query(models.Product).filter(models.Product.sku == product_in.sku).first()
        if exists_sku:
            raise HTTPException(status_code=400, detail="SKU sudah digunakan")
            
    if product_in.barcode and product_in.barcode != product.barcode:
        exists_barcode = db.query(models.Product).filter(models.Product.barcode == product_in.barcode).first()
        if exists_barcode:
            raise HTTPException(status_code=400, detail="Barcode sudah digunakan")
            
    # Check if stock is updated manually (Opname)
    stock_diff = 0.0
    if product_in.stock is not None and product_in.stock != product.stock:
        stock_diff = product_in.stock - product.stock
        
    for k, v in product_in.dict(exclude_unset=True).items():
        setattr(product, k, v)
        
    db.commit()
    db.refresh(product)
    
    # Handle manual stock adjustment movement if stock was changed
    if stock_diff != 0:
        mov = models.StockMovement(
            product_id=product.id,
            type="Penyesuaian" if stock_diff > 0 else "Penyesuaian",
            qty=abs(stock_diff),
            reference="Manual Update",
            notes=f"Penyesuaian stok manual dari {product_in.stock - stock_diff} ke {product.stock}",
            created_by=current_user.id
        )
        db.add(mov)
        db.commit()
        
    utils.check_low_stock(db, product)
    utils.create_audit_log(db, current_user.id, "Update Product", "Products", f"Memperbarui produk: {product.name}")
    
    return product

@app.delete("/api/products/{product_id}")
def delete_product(
    product_id: int, 
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "staff_gudang"]))
):
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produk tidak ditemukan")
        
    db.delete(product)
    db.commit()
    
    utils.create_audit_log(db, current_user.id, "Delete Product", "Products", f"Menghapus produk: {product.name} (SKU: {product.sku})")
    return {"message": "Produk berhasil dihapus"}

# Products EXCEL Export
@app.get("/api/products/export/excel")
def export_products_excel(db: Session = Depends(get_db)):
    products = db.query(models.Product).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Produk"
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1A365D")
    meta_font = Font(name="Calibri", size=10, italic=True, color="555555")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Title
    ws['A1'] = "LAPORAN DATA PRODUK"
    ws['A1'].font = title_font
    ws.merge_cells('A1:J1')
    
    # Metadata
    ws['A2'] = f"Unduh Pada: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    ws['A2'].font = meta_font
    ws.merge_cells('A2:J2')
    
    ws.append([]) # Spacer row
    
    # Table headers
    headers = ["No", "SKU", "Barcode", "Nama Produk", "Kategori", "Harga Beli", "Harga Jual", "Stok", "Min Stok", "Satuan"]
    ws.append(headers)
    header_row = ws.max_row
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
    ws.row_dimensions[header_row].height = 25
    
    # Data rows
    for idx, p in enumerate(products, 1):
        category_name = p.category.name if p.category else ""
        row_data = [
            idx,
            p.sku or "",
            p.barcode or "",
            p.name or "",
            category_name,
            p.cost_price,
            p.sell_price,
            p.stock,
            p.min_stock,
            p.unit or ""
        ]
        ws.append(row_data)
        curr_row = ws.max_row
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
            # Alignments & formats
            if col_idx in [1, 2, 3, 10]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in [4, 5]:
                cell.alignment = Alignment(horizontal="left")
            elif col_idx in [6, 7]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '"Rp "#,##0'
            elif col_idx in [8, 9]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
                
        # Alternate row background color
        if idx % 2 == 0:
            for col_idx in range(1, len(row_data) + 1):
                ws.cell(row=curr_row, column=col_idx).fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                
    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 4 and cell.value is not None:
                val_str = str(cell.value)
                if cell.column in [6, 7]: # Currency cols
                    val_str = f"Rp {cell.value:,.0f}"
                elif cell.column in [8, 9]: # Numeric qty cols
                    val_str = f"{cell.value:,.0f}"
                max_len = max(max_len, len(val_str))
            elif cell.row == 4 and cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    return StreamingResponse(
        file_stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=produk_warungkita.xlsx"}
    )


# ==========================================
# 7. PURCHASING MODULE (Admin, Owner, Gudang)
# ==========================================

@app.get("/api/purchases", response_model=List[schemas.PurchaseResponse])
def get_purchases(db: Session = Depends(get_db)):
    return db.query(models.Purchase).order_by(models.Purchase.purchase_date.desc()).all()

@app.get("/api/purchases/{purchase_id}", response_model=schemas.PurchaseResponse)
def get_purchase(purchase_id: int, db: Session = Depends(get_db)):
    p = db.query(models.Purchase).filter(models.Purchase.id == purchase_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Pembelian tidak ditemukan")
    return p

@app.post("/api/purchases", response_model=schemas.PurchaseResponse)
def create_purchase(
    purchase_in: schemas.PurchaseCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "staff_gudang"]))
):
    # Calculate Total
    total = 0.0
    for item in purchase_in.items:
        # Validate product
        prod = db.query(models.Product).filter(models.Product.id == item.product_id).first()
        if not prod:
            raise HTTPException(status_code=400, detail=f"Produk ID {item.product_id} tidak ditemukan")
        total += item.qty * item.cost_price
        
    # Create Purchase
    purchase = models.Purchase(
        supplier_id=purchase_in.supplier_id,
        total_amount=total,
        payment_status=purchase_in.payment_status,
        due_date=purchase_in.due_date,
        created_by=current_user.id
    )
    db.add(purchase)
    db.commit()
    db.refresh(purchase)
    
    # Add items, update stocks, log stock movements
    for item_in in purchase_in.items:
        subtotal = item_in.qty * item_in.cost_price
        item = models.PurchaseItem(
            purchase_id=purchase.id,
            product_id=item_in.product_id,
            qty=item_in.qty,
            cost_price=item_in.cost_price,
            subtotal=subtotal
        )
        db.add(item)
        
        # Increase Stock
        product = db.query(models.Product).filter(models.Product.id == item_in.product_id).first()
        product.stock += item_in.qty
        
        # Update cost_price of product (purchase price update)
        product.cost_price = item_in.cost_price
        
        # Log Stock Movement
        mov = models.StockMovement(
            product_id=product.id,
            type="Masuk",
            qty=item_in.qty,
            reference=f"Purchase ID {purchase.id}",
            notes=f"Barang masuk dari pembelian. Supplier ID {purchase.supplier_id}",
            created_by=current_user.id
        )
        db.add(mov)
        
    db.commit()
    
    # Financial Flow:
    if purchase.payment_status == "Hutang":
        # Create Debt Record
        debt = models.Debt(
            supplier_id=purchase.supplier_id,
            purchase_id=purchase.id,
            amount=total,
            paid_amount=0.0,
            status="Belum Lunas",
            due_date=purchase.due_date
        )
        db.add(debt)
        
        # Create Notification for Debt
        utils.create_notification(
            db, 
            title="Hutang Baru Tercatat",
            message=f"Hutang kepada supplier senilai Rp {total:,.0f} telah dicatat. Jatuh tempo: {purchase.due_date.strftime('%d-%m-%Y') if purchase.due_date else '-'}",
            type="Debt Due"
        )
    else:
        # Payment is Cash/Lunas, creates Cash Outflow
        cash_out = models.CashTransaction(
            type="Keluar",
            amount=total,
            category="Pembelian",
            description=f"Pembelian barang. Pembelian ID {purchase.id}",
            reference=f"Purchase ID {purchase.id}",
            created_by=current_user.id
        )
        db.add(cash_out)
        
    db.commit()
    db.refresh(purchase)
    
    # Audit log
    utils.create_audit_log(
        db, 
        current_user.id, 
        "Create Purchase", 
        "Purchasing", 
        f"Mencatat pembelian ID {purchase.id} dari Supplier ID {purchase.supplier_id} senilai Rp {total:,.0f}"
    )
    
    return purchase


# ==========================================
# 8. SALES & POS MODULE (Admin, Owner, Kasir)
# ==========================================

@app.get("/api/sales", response_model=List[schemas.SaleResponse])
def get_sales(db: Session = Depends(get_db)):
    return db.query(models.Sale).order_by(models.Sale.sale_date.desc()).all()

@app.get("/api/sales/{sale_id}", response_model=schemas.SaleResponse)
def get_sale(sale_id: int, db: Session = Depends(get_db)):
    s = db.query(models.Sale).filter(models.Sale.id == sale_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Transaksi penjualan tidak ditemukan")
    return s

@app.post("/api/sales", response_model=schemas.SaleResponse)
def create_sale(
    sale_in: schemas.SaleCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "kasir"]))
):
    # 1. Calculate Total & Validate stocks
    total = 0.0
    for item in sale_in.items:
        prod = db.query(models.Product).filter(models.Product.id == item.product_id).first()
        if not prod:
            raise HTTPException(status_code=400, detail=f"Produk ID {item.product_id} tidak ditemukan")
            
        if prod.stock < item.qty:
            raise HTTPException(
                status_code=400, 
                detail=f"Stok tidak mencukupi untuk '{prod.name}'. Stok saat ini: {prod.stock} {prod.unit}, diminta: {item.qty}"
            )
        total += item.qty * item.sell_price
        
    # Apply discount
    discount = sale_in.discount or 0.0
    final_amount = max(0.0, total - discount)
    
    # Loyalty Points calculation: 1 point for every Rp 10.000 spent
    points_earned = int(final_amount // 10000)
    
    # 2. Create Sale Record
    sale = models.Sale(
        customer_id=sale_in.customer_id,
        total_amount=total,
        discount=discount,
        final_amount=final_amount,
        payment_method=sale_in.payment_method,
        payment_status=sale_in.payment_status,
        due_date=sale_in.due_date,
        points_earned=points_earned,
        created_by=current_user.id
    )
    db.add(sale)
    db.commit()
    db.refresh(sale)
    
    # 3. Add Sale Items, decrease stock, log stock movement, check low stock
    for item_in in sale_in.items:
        subtotal = item_in.qty * item_in.sell_price
        item = models.SaleItem(
            sale_id=sale.id,
            product_id=item_in.product_id,
            qty=item_in.qty,
            sell_price=item_in.sell_price,
            subtotal=subtotal
        )
        db.add(item)
        
        # Decrease Stock
        product = db.query(models.Product).filter(models.Product.id == item_in.product_id).first()
        product.stock -= item_in.qty
        
        # Stock movement log
        mov = models.StockMovement(
            product_id=product.id,
            type="Keluar",
            qty=item_in.qty,
            reference=f"Sale ID {sale.id}",
            notes=f"Barang keluar dari penjualan POS. Invoice ID {sale.id}",
            created_by=current_user.id
        )
        db.add(mov)
        
        # Check for stock alerts
        utils.check_low_stock(db, product)
        
    db.commit()
    
    # 4. Handle loyalty points & total transaction updates for customer
    if sale.customer_id:
        customer = db.query(models.Customer).filter(models.Customer.id == sale.customer_id).first()
        if customer:
            customer.total_transactions += final_amount
            # Skip Pelanggan Umum ID = 1 points
            if customer.id != 1:
                customer.points += points_earned
                
    db.commit()
    
    # 5. Financial Flow:
    if sale.payment_status == "Piutang":
        # Customer buys on credit. Create Receivable Record
        if not sale.customer_id:
            raise HTTPException(status_code=400, detail="Pelanggan harus dipilih jika melakukan transaksi piutang")
            
        receivable = models.Receivable(
            customer_id=sale.customer_id,
            sale_id=sale.id,
            amount=final_amount,
            paid_amount=0.0,
            status="Belum Lunas",
            due_date=sale.due_date
        )
        db.add(receivable)
        
        utils.create_notification(
            db,
            title="Piutang Baru Tercatat",
            message=f"Piutang pelanggan senilai Rp {final_amount:,.0f} telah dicatat. Jatuh tempo: {sale.due_date.strftime('%d-%m-%Y') if sale.due_date else '-'}",
            type="Receivables Due"
        )
    else:
        # Cash/Lunas. Add Cash Inflow
        cash_in = models.CashTransaction(
            type="Masuk",
            amount=final_amount,
            category="Penjualan",
            description=f"Penjualan barang POS. Invoice ID {sale.id}. Metode: {sale.payment_method}",
            reference=f"Sale ID {sale.id}",
            created_by=current_user.id
        )
        db.add(cash_in)
        
    db.commit()
    db.refresh(sale)
    
    # Audit log
    utils.create_audit_log(
        db, 
        current_user.id, 
        "Create Sale", 
        "Sales", 
        f"Mencatat transaksi penjualan POS ID {sale.id} senilai Rp {final_amount:,.0f} via {sale.payment_method}"
    )
    
    return sale

# POS Receipt PDF Export
@app.get("/api/sales/{sale_id}/pdf")
def get_receipt_pdf(sale_id: int, db: Session = Depends(get_db)):
    sale = db.query(models.Sale).filter(models.Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
        
    buffer = io.BytesIO()
    
    # Generate receipt PDF layout
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Header
    p.setFont("Helvetica-Bold", 18)
    p.setFillColor(colors.HexColor("#1A365D")) # Slate Blue
    p.drawString(100, height - 80, "WARUNGKITA")
    p.setFont("Helvetica", 10)
    p.setFillColor(colors.HexColor("#4A5568"))
    p.drawString(100, height - 95, "Jl. Raya Warung Kemitraan No. 8, Jakarta")
    p.drawString(100, height - 110, "Telepon: 0812-3456-7890")
    
    # Receipt details
    p.setFont("Helvetica-Bold", 12)
    p.drawString(100, height - 150, "FAKTUR PENJUALAN")
    p.setFont("Helvetica", 10)
    p.drawString(100, height - 170, f"No. Transaksi : POS-{sale.id}")
    p.drawString(100, height - 185, f"Tanggal       : {sale.sale_date.strftime('%d-%m-%Y %H:%M:%S')}")
    p.drawString(100, height - 200, f"Kasir         : {sale.created_by}")
    p.drawString(100, height - 215, f"Pelanggan     : {sale.customer.name if sale.customer else 'Pelanggan Umum'}")
    p.drawString(100, height - 230, f"Pembayaran    : {sale.payment_method} ({sale.payment_status})")
    
    # Divider
    p.setStrokeColor(colors.HexColor("#E2E8F0"))
    p.line(100, height - 245, width - 100, height - 245)
    
    # Items table headers
    p.setFont("Helvetica-Bold", 10)
    p.drawString(100, height - 260, "Nama Produk")
    p.drawString(300, height - 260, "Qty")
    p.drawString(350, height - 260, "Harga")
    p.drawString(450, height - 260, "Subtotal")
    p.line(100, height - 268, width - 100, height - 268)
    
    y = height - 285
    p.setFont("Helvetica", 10)
    for item in sale.items:
        if y < 100:
            p.showPage()
            y = height - 80
            p.setFont("Helvetica", 10)
            
        p.drawString(100, y, item.product.name[:35] if item.product else f"Produk ID {item.product_id}")
        p.drawString(300, y, f"{item.qty:.0f}")
        p.drawString(350, y, f"Rp {item.sell_price:,.0f}")
        p.drawString(450, y, f"Rp {item.subtotal:,.0f}")
        y -= 20
        
    p.line(100, y - 5, width - 100, y - 5)
    y -= 25
    
    # Total summary
    p.setFont("Helvetica", 10)
    p.drawString(300, y, "Total Belanja :")
    p.drawRightString(width - 100, y, f"Rp {sale.total_amount:,.0f}")
    y -= 15
    p.drawString(300, y, "Diskon        :")
    p.drawRightString(width - 100, y, f"Rp {sale.discount:,.0f}")
    y -= 20
    
    p.setFont("Helvetica-Bold", 11)
    p.drawString(300, y, "Total Akhir   :")
    p.drawRightString(width - 100, y, f"Rp {sale.final_amount:,.0f}")
    y -= 30
    
    # Thank you footer
    p.setFont("Helvetica-Oblique", 9)
    p.setFillColor(colors.HexColor("#718096"))
    p.drawCentredString(width / 2.0, y - 40, "Terima Kasih atas Kunjungan Anda!")
    p.drawCentredString(width / 2.0, y - 55, "Barang yang sudah dibeli tidak dapat ditukar/dikembalikan")
    
    p.showPage()
    p.save()
    buffer.seek(0)
    
    return StreamingResponse(
        buffer, 
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=struk_pos_{sale.id}.pdf"}
    )


# ==========================================
# 9. STOCK MODULE (Admin, Owner, Gudang)
# ==========================================

@app.get("/api/stock/movements", response_model=List[schemas.StockMovementResponse])
def get_stock_movements(db: Session = Depends(get_db)):
    return db.query(models.StockMovement).order_by(models.StockMovement.created_at.desc()).all()

from pydantic import BaseModel as PydanticBaseModel

class StockAdjustFrontend(PydanticBaseModel):
    product_id: int
    qty: int
    type: str
    reason: str = "Stockopname"

@app.post("/api/stock/adjust")
def adjust_stock(
    adj: StockAdjustFrontend,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "staff_gudang"]))
):
    product = db.query(models.Product).filter(models.Product.id == adj.product_id).first()
    if not product:
        raise HTTPException(status_code=404, detail="Produk tidak ditemukan")

    is_in = adj.qty > 0
    abs_qty = abs(adj.qty)

    if not is_in and product.stock < abs_qty:
        raise HTTPException(status_code=400, detail=f"Stok saat ini ({product.stock}) tidak mencukupi.")

    product.stock += adj.qty

    movement = models.StockMovement(
        product_id=product.id,
        type=adj.type,
        qty=adj.qty,
        reference="Manual Adjustment",
        notes=adj.reason,
        created_by=current_user.id
    )
    db.add(movement)
    db.commit()
    db.refresh(movement)
    utils.check_low_stock(db, product)
    utils.create_audit_log(db, current_user.id, "Adjust Stock", "Stock",
                           f"Penyesuaian stok '{product.name}': {adj.qty:+} ({adj.reason}). Stok akhir: {product.stock}")
    return {
        "id": movement.id,
        "product_id": movement.product_id,
        "qty": movement.qty,
        "type": movement.type,
        "reason": movement.notes,
        "created_at": movement.created_at.isoformat() if movement.created_at else "",
        "product": {"name": product.name, "sku": product.sku},
        "user": {"full_name": current_user.full_name}
    }


# ==========================================
# 10. CASH MODULE (Admin, Owner, Kasir)
# ==========================================

@app.get("/api/cash/transactions", response_model=List[schemas.CashTransactionResponse])
def get_cash_transactions(db: Session = Depends(get_db)):
    return db.query(models.CashTransaction).order_by(models.CashTransaction.created_at.desc()).all()

@app.get("/api/cash/balance")
def get_cash_balance(db: Session = Depends(get_db)):
    # Calculate sum of Inflows and Outflows
    cash_in = db.query(func.sum(models.CashTransaction.amount)).filter(models.CashTransaction.type == "Masuk").scalar() or 0.0
    cash_out = db.query(func.sum(models.CashTransaction.amount)).filter(models.CashTransaction.type == "Keluar").scalar() or 0.0
    balance = cash_in - cash_out
    return {"balance": balance, "total_cash_in": cash_in, "total_cash_out": cash_out}

@app.post("/api/cash/transactions", response_model=schemas.CashTransactionResponse)
def create_cash_transaction(
    tx: schemas.CashTransactionCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "kasir"]))
):
    if tx.type not in ["Masuk", "Keluar"]:
        raise HTTPException(status_code=400, detail="Tipe transaksi kas harus Masuk atau Keluar")
        
    cash_tx = models.CashTransaction(
        type=tx.type,
        amount=tx.amount,
        category=tx.category,
        description=tx.description,
        reference="Manual Input",
        created_by=current_user.id
    )
    db.add(cash_tx)
    db.commit()
    db.refresh(cash_tx)
    
    utils.create_audit_log(
        db, 
        current_user.id, 
        "Create Cash Transaction", 
        "Cash", 
        f"Mencatat kas {tx.type} senilai Rp {tx.amount:,.0f} kategori {tx.category}"
    )
    
    return cash_tx


# ==========================================
# 11. DEBT MODULE (Admin, Owner)
# ==========================================

class DebtPayFrontend(PydanticBaseModel):
    amount: float
    notes: str = ""

@app.get("/api/debts")
def get_debts(db: Session = Depends(get_db)):
    debts = db.query(models.Debt).order_by(models.Debt.due_date.asc()).all()
    result = []
    for d in debts:
        supplier = db.query(models.Supplier).filter(models.Supplier.id == d.supplier_id).first()
        result.append({
            "id": d.id,
            "supplier_id": d.supplier_id,
            "purchase_id": d.purchase_id,
            "total_amount": d.amount,
            "remaining_amount": d.amount - d.paid_amount,
            "status": d.status,
            "due_date": d.due_date.isoformat() if d.due_date else None,
            "supplier": {"name": supplier.name} if supplier else None,
            "payments": []
        })
    return result

@app.post("/api/debts/{debt_id}/pay")
def pay_debt(
    debt_id: int,
    payment: DebtPayFrontend,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner"]))
):
    debt = db.query(models.Debt).filter(models.Debt.id == debt_id).first()
    if not debt:
        raise HTTPException(status_code=404, detail="Hutang tidak ditemukan")
    remaining = debt.amount - debt.paid_amount
    if payment.amount > remaining:
        raise HTTPException(status_code=400, detail=f"Jumlah pembayaran melebihi sisa hutang (Sisa: Rp {remaining:,.0f})")
    debt.paid_amount += payment.amount
    if debt.paid_amount >= debt.amount:
        debt.status = "Lunas"
        if debt.purchase:
            debt.purchase.payment_status = "Lunas"
    cash_out = models.CashTransaction(
        type="Keluar", amount=payment.amount,
        category="Pembayaran Hutang",
        description=payment.notes or f"Pelunasan hutang ID {debt.id}",
        reference=f"Debt ID {debt.id}", created_by=current_user.id
    )
    db.add(cash_out)
    db.commit()
    db.refresh(debt)
    utils.create_audit_log(db, current_user.id, "Pay Debt", "Debts",
                           f"Membayar hutang ID {debt.id} Rp {payment.amount:,.0f}")
    supplier = db.query(models.Supplier).filter(models.Supplier.id == debt.supplier_id).first()
    return {
        "id": debt.id,
        "supplier_id": debt.supplier_id,
        "purchase_id": debt.purchase_id,
        "total_amount": debt.amount,
        "remaining_amount": debt.amount - debt.paid_amount,
        "status": debt.status,
        "due_date": debt.due_date.isoformat() if debt.due_date else None,
        "supplier": {"name": supplier.name} if supplier else None,
        "payments": []
    }


# ==========================================
# 12. RECEIVABLES MODULE (Admin, Owner, Kasir)
# ==========================================

class ReceivablePayFrontend(PydanticBaseModel):
    amount: float
    notes: str = ""

@app.get("/api/receivables")
def get_receivables(db: Session = Depends(get_db)):
    receivables = db.query(models.Receivable).order_by(models.Receivable.due_date.asc()).all()
    result = []
    for r in receivables:
        customer = db.query(models.Customer).filter(models.Customer.id == r.customer_id).first()
        result.append({
            "id": r.id,
            "customer_id": r.customer_id,
            "sale_id": r.sale_id,
            "total_amount": r.amount,
            "remaining_amount": r.amount - r.paid_amount,
            "status": r.status,
            "due_date": r.due_date.isoformat() if r.due_date else None,
            "customer": {"name": customer.name} if customer else None,
            "payments": []
        })
    return result

@app.post("/api/receivables/{receivable_id}/pay")
def pay_receivable(
    receivable_id: int,
    payment: ReceivablePayFrontend,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "kasir"]))
):
    receivable = db.query(models.Receivable).filter(models.Receivable.id == receivable_id).first()
    if not receivable:
        raise HTTPException(status_code=404, detail="Piutang tidak ditemukan")
    remaining = receivable.amount - receivable.paid_amount
    if payment.amount > remaining:
        raise HTTPException(status_code=400, detail=f"Melebihi sisa piutang (Rp {remaining:,.0f})")
    receivable.paid_amount += payment.amount
    if receivable.paid_amount >= receivable.amount:
        receivable.status = "Lunas"
        if receivable.sale:
            receivable.sale.payment_status = "Lunas"
    cash_in = models.CashTransaction(
        type="Masuk", amount=payment.amount,
        category="Pelunasan Piutang",
        description=payment.notes or f"Pelunasan piutang ID {receivable.id}",
        reference=f"Receivable ID {receivable.id}", created_by=current_user.id
    )
    db.add(cash_in)
    db.commit()
    db.refresh(receivable)
    utils.create_audit_log(db, current_user.id, "Pay Receivable", "Receivables",
                           f"Menerima piutang ID {receivable.id} Rp {payment.amount:,.0f}")
    customer = db.query(models.Customer).filter(models.Customer.id == receivable.customer_id).first()
    return {
        "id": receivable.id,
        "customer_id": receivable.customer_id,
        "sale_id": receivable.sale_id,
        "total_amount": receivable.amount,
        "remaining_amount": receivable.amount - receivable.paid_amount,
        "status": receivable.status,
        "due_date": receivable.due_date.isoformat() if receivable.due_date else None,
        "customer": {"name": customer.name} if customer else None,
        "payments": []
    }


# ==========================================
# 13. NOTIFICATIONS MODULE
# ==========================================

@app.get("/api/notifications", response_model=List[schemas.NotificationResponse])
def get_notifications(db: Session = Depends(get_db)):
    return db.query(models.Notification).filter(models.Notification.is_read == False).order_by(models.Notification.created_at.desc()).all()

@app.post("/api/notifications/{notification_id}/read")
def read_notification(notification_id: int, db: Session = Depends(get_db)):
    notif = db.query(models.Notification).filter(models.Notification.id == notification_id).first()
    if not notif:
        raise HTTPException(status_code=404, detail="Notifikasi tidak ditemukan")
        
    notif.is_read = True
    db.commit()
    return {"message": "Notifikasi ditandai sebagai dibaca"}


# ==========================================
# 14. AUDIT LOGS MODULE (Admin only)
# ==========================================

@app.get("/api/audit-logs", response_model=List[schemas.AuditLogResponse])
def get_audit_logs(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin"]))
):
    return db.query(models.AuditLog).order_by(models.AuditLog.created_at.desc()).all()


# ==========================================
# 15. REPORTS & DASHBOARD MODULE
# ==========================================

@app.get("/api/reports/dashboard")
def get_dashboard_analytics(db: Session = Depends(get_db)):
    today = datetime.date.today()
    today_start = datetime.datetime.combine(today, datetime.time.min)
    today_end = datetime.datetime.combine(today, datetime.time.max)
    
    month_start = datetime.datetime(today.year, today.month, 1)
    
    # 1. Omzet Hari Ini (Total final_amount of sales today)
    sales_today = db.query(func.sum(models.Sale.final_amount)).filter(
        and_(models.Sale.sale_date >= today_start, models.Sale.sale_date <= today_end)
    ).scalar() or 0.0
    
    # 2. Transaksi Hari Ini
    tx_today = db.query(func.count(models.Sale.id)).filter(
        and_(models.Sale.sale_date >= today_start, models.Sale.sale_date <= today_end)
    ).scalar() or 0
    
    # 3. Omzet Bulan Ini
    sales_month = db.query(func.sum(models.Sale.final_amount)).filter(
        models.Sale.sale_date >= month_start
    ).scalar() or 0.0
    
    # 4. Profit Bulan Ini (Laba Kotor = Penjualan - Harga Pokok Penjualan)
    # HPP = sum(qty * cost_price of products sold this month)
    sales_items_month = db.query(
        models.SaleItem.qty, 
        models.Product.cost_price, 
        models.SaleItem.sell_price
    ).join(models.Sale).join(models.Product).filter(
        models.Sale.sale_date >= month_start
    ).all()
    
    profit_month = 0.0
    for qty, cost, sell in sales_items_month:
        profit_month += qty * (sell - cost)
        
    # Subtract discount applied in sales (approximate proportionate discount)
    total_sales_month_raw = db.query(func.sum(models.Sale.total_amount)).filter(
        models.Sale.sale_date >= month_start
    ).scalar() or 1.0
    total_discount_month = db.query(func.sum(models.Sale.discount)).filter(
        models.Sale.sale_date >= month_start
    ).scalar() or 0.0
    
    profit_month = max(0.0, profit_month - total_discount_month)
    
    # 5. Produk Terlaris (Top 5 products by quantity sold)
    top_products_query = db.query(
        models.Product.name,
        func.sum(models.SaleItem.qty).label("sold_qty")
    ).join(models.SaleItem).group_by(models.Product.id).order_by(func.sum(models.SaleItem.qty).desc()).limit(5).all()
    
    top_products = [{"name": name, "qty": float(qty)} for name, qty in top_products_query]
    
    # 6. Produk Stok Hampir Habis (stok <= min_stock)
    low_stock_count = db.query(func.count(models.Product.id)).filter(
        models.Product.stock <= models.Product.min_stock
    ).scalar() or 0
    
    low_stock_items = db.query(models.Product).filter(
        models.Product.stock <= models.Product.min_stock
    ).limit(5).all()
    
    low_stock_list = [{"sku": p.sku, "name": p.name, "stock": p.stock, "unit": p.unit} for p in low_stock_items]
    
    # 7. Top Customers
    top_customers_query = db.query(
        models.Customer.name,
        models.Customer.total_transactions
    ).filter(models.Customer.id != 1).order_by(models.Customer.total_transactions.desc()).limit(5).all()
    
    top_customers = [{"name": name, "spent": spent} for name, spent in top_customers_query]
    
    # 8. Sales Chart (Last 7 Days)
    chart_data = []
    for i in range(6, -1, -1):
        day = today - datetime.timedelta(days=i)
        d_start = datetime.datetime.combine(day, datetime.time.min)
        d_end = datetime.datetime.combine(day, datetime.time.max)
        
        day_sales = db.query(func.sum(models.Sale.final_amount)).filter(
            and_(models.Sale.sale_date >= d_start, models.Sale.sale_date <= d_end)
        ).scalar() or 0.0
        
        chart_data.append({
            "date": day.strftime("%A"),
            "sales": day_sales
        })
        
    return {
        "omzet_today": sales_today,
        "tx_today": tx_today,
        "omzet_month": sales_month,
        "profit_month": profit_month,
        "top_products": top_products,
        "low_stock_count": low_stock_count,
        "low_stock_list": low_stock_list,
        "top_customers": top_customers,
        "chart_data": chart_data
    }

@app.get("/api/reports/sales-report")
def get_sales_report(
    db: Session = Depends(get_db),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    query = db.query(models.Sale)
    if start_date:
        query = query.filter(models.Sale.sale_date >= datetime.datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        query = query.filter(models.Sale.sale_date <= datetime.datetime.strptime(end_date, "%Y-%m-%d") + datetime.timedelta(days=1))
        
    sales = query.order_by(models.Sale.sale_date.desc()).all()
    
    total_revenue = sum(s.final_amount for s in sales)
    total_sales = len(sales)
    total_discount = sum(s.discount for s in sales)
    
    return {
        "sales": sales,
        "total_revenue": total_revenue,
        "total_sales": total_sales,
        "total_discount": total_discount
    }

@app.get("/api/reports/purchases-report")
def get_purchases_report(
    db: Session = Depends(get_db),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    query = db.query(models.Purchase)
    if start_date:
        query = query.filter(models.Purchase.purchase_date >= datetime.datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        query = query.filter(models.Purchase.purchase_date <= datetime.datetime.strptime(end_date, "%Y-%m-%d") + datetime.timedelta(days=1))
        
    purchases = query.order_by(models.Purchase.purchase_date.desc()).all()
    total_spending = sum(p.total_amount for p in purchases)
    
    return {
        "purchases": purchases,
        "total_spending": total_spending
    }

@app.get("/api/reports/profit-loss")
def get_profit_loss_report(
    db: Session = Depends(get_db),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    sales_q = db.query(func.sum(models.Sale.final_amount))
    hpp_q = db.query(func.sum(models.SaleItem.qty * models.Product.cost_price)).join(models.Sale).join(models.Product)
    op_q = db.query(func.sum(models.CashTransaction.amount)).filter(
        and_(models.CashTransaction.type == "Keluar")
    )

    if start_date:
        try:
            if "T" in start_date:
                dt_s = datetime.datetime.fromisoformat(start_date.replace("Z", "+00:00"))
            else:
                dt_s = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            sales_q = sales_q.filter(models.Sale.sale_date >= dt_s)
            hpp_q = hpp_q.filter(models.Sale.sale_date >= dt_s)
            op_q = op_q.filter(models.CashTransaction.created_at >= dt_s)
        except Exception:
            pass
    if end_date:
        try:
            if "T" in end_date:
                dt_e = datetime.datetime.fromisoformat(end_date.replace("Z", "+00:00"))
            else:
                dt_e = datetime.datetime.strptime(end_date, "%Y-%m-%d") + datetime.timedelta(days=1)
            sales_q = sales_q.filter(models.Sale.sale_date <= dt_e)
            hpp_q = hpp_q.filter(models.Sale.sale_date <= dt_e)
            op_q = op_q.filter(models.CashTransaction.created_at <= dt_e)
        except Exception:
            pass

    revenue = sales_q.scalar() or 0.0
    cogs = hpp_q.scalar() or 0.0
    expenses = op_q.scalar() or 0.0
    gross_profit = revenue - cogs
    net_profit = gross_profit - expenses

    return {
        "revenue": revenue,
        "cogs": cogs,
        "gross_profit": gross_profit,
        "expenses": expenses,
        "net_profit": net_profit
    }


# ==========================================
# ADAPTER / ALIAS ENDPOINTS
# (maps frontend API paths to backend logic)
# ==========================================

# Cash: /api/cash/summary — returns balance + transaction logs in one response
@app.get("/api/cash/summary")
def get_cash_summary(db: Session = Depends(get_db)):
    cash_in = db.query(func.sum(models.CashTransaction.amount)).filter(models.CashTransaction.type == "Masuk").scalar() or 0.0
    cash_out = db.query(func.sum(models.CashTransaction.amount)).filter(models.CashTransaction.type == "Keluar").scalar() or 0.0
    balance = cash_in - cash_out
    logs = db.query(models.CashTransaction).order_by(models.CashTransaction.created_at.desc()).all()

    def tx_to_dict(tx):
        return {
            "id": tx.id,
            "tx_type": tx.type,
            "amount": tx.amount,
            "category": tx.category,
            "notes": tx.description or "",
            "tx_date": tx.created_at.isoformat() if tx.created_at else "",
            "reference_id": None
        }

    return {
        "inflow": cash_in,
        "outflow": cash_out,
        "balance": balance,
        "logs": [tx_to_dict(t) for t in logs]
    }

# Cash: POST /api/cash/transaction — maps frontend field names to CashTransaction model
class CashTransactionFrontendIn(schemas.BaseSchema if hasattr(schemas, "BaseSchema") else object):
    pass

from pydantic import BaseModel as PydanticBaseModel

class CashTransactionFrontendCreate(PydanticBaseModel):
    tx_type: str
    amount: float
    category: str
    notes: str = ""

@app.post("/api/cash/transaction")
def create_cash_transaction_alias(
    tx: CashTransactionFrontendCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin", "owner", "kasir"]))
):
    mapped_type = "Masuk" if tx.tx_type == "In" else "Keluar"
    cash_tx = models.CashTransaction(
        type=mapped_type,
        amount=tx.amount,
        category=tx.category,
        description=tx.notes,
        reference="Manual Input",
        created_by=current_user.id
    )
    db.add(cash_tx)
    db.commit()
    db.refresh(cash_tx)
    utils.create_audit_log(db, current_user.id, "Cash Transaction", "Cash",
                           f"Kas {mapped_type} Rp {tx.amount:,.0f} - {tx.category}")
    return {"id": cash_tx.id, "message": "Transaksi kas berhasil dicatat"}

# Audit: /api/audit/logs — alias for /api/audit-logs
@app.get("/api/audit/logs")
def get_audit_logs_alias(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.RoleChecker(["super_admin"]))
):
    logs = db.query(models.AuditLog).order_by(models.AuditLog.created_at.desc()).all()
    result = []
    for log in logs:
        user = db.query(models.User).filter(models.User.id == log.user_id).first()
        result.append({
            "id": log.id,
            "user_id": log.user_id,
            "action": log.action,
            "details": log.details,
            "ip_address": "127.0.0.1",
            "created_at": log.created_at.isoformat() if log.created_at else "",
            "user": {
                "username": user.username if user else "system",
                "full_name": user.full_name if user else "System",
                "role": user.role if user else "system"
            } if user else None
        })
    return result

# Reports: /api/reports/sales — alias with ISO date filter params
@app.get("/api/reports/sales")
def get_sales_report_alias(
    db: Session = Depends(get_db),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    query = db.query(models.Sale)
    if start_date:
        try:
            query = query.filter(models.Sale.sale_date >= datetime.datetime.fromisoformat(start_date.replace("Z", "+00:00")))
        except Exception:
            pass
    if end_date:
        try:
            query = query.filter(models.Sale.sale_date <= datetime.datetime.fromisoformat(end_date.replace("Z", "+00:00")))
        except Exception:
            pass

    sales = query.order_by(models.Sale.sale_date.desc()).all()
    result = []
    for s in sales:
        customer = db.query(models.Customer).filter(models.Customer.id == s.customer_id).first() if s.customer_id else None
        result.append({
            "id": s.id,
            "sale_date": s.sale_date.isoformat() if s.sale_date else "",
            "total_amount": s.total_amount,
            "discount": s.discount,
            "final_amount": s.final_amount,
            "payment_method": s.payment_method,
            "payment_status": s.payment_status,
            "customer": {"name": customer.name} if customer else None
        })
    return result

# Reports: /api/reports/purchases — alias with ISO date filter
@app.get("/api/reports/purchases")
def get_purchases_report_alias(
    db: Session = Depends(get_db),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    query = db.query(models.Purchase)
    if start_date:
        try:
            query = query.filter(models.Purchase.purchase_date >= datetime.datetime.fromisoformat(start_date.replace("Z", "+00:00")))
        except Exception:
            pass
    if end_date:
        try:
            query = query.filter(models.Purchase.purchase_date <= datetime.datetime.fromisoformat(end_date.replace("Z", "+00:00")))
        except Exception:
            pass

    purchases = query.order_by(models.Purchase.purchase_date.desc()).all()
    result = []
    for p in purchases:
        supplier = db.query(models.Supplier).filter(models.Supplier.id == p.supplier_id).first() if p.supplier_id else None
        result.append({
            "id": p.id,
            "purchase_date": p.purchase_date.isoformat() if p.purchase_date else "",
            "total_amount": p.total_amount,
            "payment_status": p.payment_status,
            "supplier": {"name": supplier.name} if supplier else None
        })
    return result

# Cleaned up alias endpoints that are now integrated natively.


# ==========================================
# 16. EXPORT REPORTS MODULE (Excel & PDF)
# ==========================================

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def add_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.HexColor("#718096"))
    canvas.setStrokeColor(colors.HexColor("#E2E8F0"))
    canvas.setLineWidth(0.5)
    canvas.line(40, 45, 572, 45) # 612 - 40 = 572
    canvas.drawString(40, 32, "WarungKita - Laporan Bisnis")
    canvas.drawRightString(572, 32, f"Halaman {doc.page}")
    canvas.restoreState()

@app.get("/api/reports/sales/export/excel")
def export_sales_report_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.Sale)
    if start_date:
        try:
            query = query.filter(models.Sale.sale_date >= datetime.datetime.fromisoformat(start_date.replace("Z", "+00:00")))
        except Exception:
            pass
    if end_date:
        try:
            query = query.filter(models.Sale.sale_date <= datetime.datetime.fromisoformat(end_date.replace("Z", "+00:00")))
        except Exception:
            pass

    sales = query.order_by(models.Sale.sale_date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"
    ws.views.sheetView[0].showGridLines = True

    title_font = Font(name="Calibri", size=16, bold=True, color="1A365D")
    meta_font = Font(name="Calibri", size=10, italic=True, color="555555")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Calibri", size=11)
    total_font = Font(name="Calibri", size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    summary_border = Border(
        top=Side(style='thin', color='1A365D'),
        bottom=Side(style='double', color='1A365D')
    )

    ws['A1'] = "LAPORAN PENJUALAN"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    
    date_str = "Semua Periode"
    if start_date and end_date:
        sd = start_date.split("T")[0]
        ed = end_date.split("T")[0]
        date_str = f"Periode: {sd} s/d {ed}"
    elif start_date:
        sd = start_date.split("T")[0]
        date_str = f"Periode: Mulai {sd}"
    elif end_date:
        ed = end_date.split("T")[0]
        date_str = f"Periode: Sampai {ed}"
        
    ws['A2'] = date_str
    ws['A2'].font = meta_font
    ws.merge_cells('A2:G2')
    
    ws['A3'] = f"Unduh Pada: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    ws['A3'].font = meta_font
    ws.merge_cells('A3:G3')
    
    ws.append([]) # Spacer

    headers = ["No. Transaksi", "Tanggal & Waktu", "Pelanggan", "Metode Pembayaran", "Status", "Diskon", "Total Akhir"]
    ws.append(headers)
    header_row = ws.max_row
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
    ws.row_dimensions[header_row].height = 25

    total_discount = 0.0
    total_amount = 0.0

    for s in sales:
        cust_name = s.customer.name if s.customer else "Pelanggan Umum"
        dt_local = s.sale_date.strftime("%d-%m-%Y %H:%M:%S") if s.sale_date else ""
        row_data = [
            f"POS-{s.id}",
            dt_local,
            cust_name,
            s.payment_method,
            s.payment_status,
            s.discount,
            s.final_amount
        ]
        ws.append(row_data)
        curr_row = ws.max_row
        
        total_discount += s.discount
        total_amount += s.final_amount
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
            if col_idx in [1, 2, 4, 5]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="left")
            elif col_idx in [6, 7]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '"Rp "#,##0'

    ws.append([]) # Spacer row
    summary_row = ws.max_row + 1
    
    ws.cell(row=summary_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=5)
    
    ws.cell(row=summary_row, column=6, value=total_discount).font = total_font
    ws.cell(row=summary_row, column=6).number_format = '"Rp "#,##0'
    ws.cell(row=summary_row, column=6).alignment = Alignment(horizontal="right")
    
    ws.cell(row=summary_row, column=7, value=total_amount).font = total_font
    ws.cell(row=summary_row, column=7).number_format = '"Rp "#,##0'
    ws.cell(row=summary_row, column=7).alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.border = summary_border
        
    ws.row_dimensions[summary_row].height = 20

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row > 4 and cell.value:
                val_str = str(cell.value)
                if isinstance(cell.value, (int, float)):
                    val_str = f"Rp {cell.value:,.0f}"
                max_len = max(max_len, len(val_str))
            elif cell.row == 4 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
                
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    fn_date = datetime.datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="laporan_penjualan_{fn_date}.xlsx"'}
    )

@app.get("/api/reports/sales/export/pdf")
def export_sales_report_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.Sale)
    if start_date:
        try:
            query = query.filter(models.Sale.sale_date >= datetime.datetime.fromisoformat(start_date.replace("Z", "+00:00")))
        except Exception:
            pass
    if end_date:
        try:
            query = query.filter(models.Sale.sale_date <= datetime.datetime.fromisoformat(end_date.replace("Z", "+00:00")))
        except Exception:
            pass

    sales = query.order_by(models.Sale.sale_date.desc()).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=60
    )

    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#1A365D"),
        fontName="Helvetica-Bold",
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'ReportSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#4A5568"),
        fontName="Helvetica",
        spaceAfter=15
    )
    th_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=colors.white,
        fontName="Helvetica-Bold",
        alignment=1
    )
    td_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#2D3748"),
        fontName="Helvetica"
    )
    td_center_style = ParagraphStyle(
        'TableCellCenter',
        parent=td_style,
        alignment=1
    )
    td_right_style = ParagraphStyle(
        'TableCellRight',
        parent=td_style,
        alignment=2
    )

    story = []

    story.append(Paragraph("LAPORAN PENJUALAN", title_style))
    
    date_str = "Semua Periode"
    if start_date and end_date:
        sd = start_date.split("T")[0]
        ed = end_date.split("T")[0]
        date_str = f"Periode: {sd} s/d {ed}"
    elif start_date:
        sd = start_date.split("T")[0]
        date_str = f"Periode: Mulai {sd}"
    elif end_date:
        ed = end_date.split("T")[0]
        date_str = f"Periode: Sampai {ed}"
        
    meta_text = f"<b>{date_str}</b> &nbsp;&nbsp;|&nbsp;&nbsp; Diunduh: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    story.append(Paragraph(meta_text, subtitle_style))

    table_data = [[
        Paragraph("No. Transaksi", th_style),
        Paragraph("Tanggal & Waktu", th_style),
        Paragraph("Pelanggan", th_style),
        Paragraph("Metode", th_style),
        Paragraph("Status", th_style),
        Paragraph("Diskon", th_style),
        Paragraph("Total Akhir", th_style)
    ]]

    total_discount = 0.0
    total_amount = 0.0

    for s in sales:
        cust_name = s.customer.name if s.customer else "Pelanggan Umum"
        dt_str = s.sale_date.strftime("%d-%m-%Y %H:%M:%S") if s.sale_date else ""
        table_data.append([
            Paragraph(f"POS-{s.id}", td_center_style),
            Paragraph(dt_str, td_center_style),
            Paragraph(cust_name, td_style),
            Paragraph(s.payment_method or "", td_center_style),
            Paragraph(s.payment_status or "", td_center_style),
            Paragraph(f"Rp {s.discount:,.0f}" if s.discount else "-", td_right_style),
            Paragraph(f"Rp {s.final_amount:,.0f}", td_right_style)
        ])
        total_discount += s.discount
        total_amount += s.final_amount

    summary_idx = len(table_data)
    table_data.append([
        Paragraph("<b>TOTAL</b>", td_style),
        "", "", "", "",
        Paragraph(f"<b>Rp {total_discount:,.0f}</b>", td_right_style),
        Paragraph(f"<b>Rp {total_amount:,.0f}</b>", td_right_style)
    ])

    t = Table(table_data, colWidths=[60, 105, 115, 60, 52, 65, 75])
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1A365D")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ])
    
    for i in range(1, summary_idx):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F7FAFC"))
            
    t_style.add('SPAN', (0, summary_idx), (4, summary_idx))
    t_style.add('BACKGROUND', (0, summary_idx), (-1, summary_idx), colors.HexColor("#EDF2F7"))
    t_style.add('LINEABOVE', (0, summary_idx), (-1, summary_idx), 1, colors.HexColor("#1A365D"))
    t_style.add('LINEBELOW', (0, summary_idx), (-1, summary_idx), 1.5, colors.HexColor("#1A365D"))

    t.setStyle(t_style)
    story.append(t)

    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    buffer.seek(0)
    
    fn_date = datetime.datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="laporan_penjualan_{fn_date}.pdf"'}
    )

@app.get("/api/reports/purchases/export/excel")
def export_purchases_report_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.Purchase)
    if start_date:
        try:
            query = query.filter(models.Purchase.purchase_date >= datetime.datetime.fromisoformat(start_date.replace("Z", "+00:00")))
        except Exception:
            pass
    if end_date:
        try:
            query = query.filter(models.Purchase.purchase_date <= datetime.datetime.fromisoformat(end_date.replace("Z", "+00:00")))
        except Exception:
            pass

    purchases = query.order_by(models.Purchase.purchase_date.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Pembelian"
    ws.views.sheetView[0].showGridLines = True

    title_font = Font(name="Calibri", size=16, bold=True, color="1A365D")
    meta_font = Font(name="Calibri", size=10, italic=True, color="555555")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Calibri", size=11)
    total_font = Font(name="Calibri", size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    summary_border = Border(
        top=Side(style='thin', color='1A365D'),
        bottom=Side(style='double', color='1A365D')
    )

    ws['A1'] = "LAPORAN PEMBELIAN"
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    
    date_str = "Semua Periode"
    if start_date and end_date:
        sd = start_date.split("T")[0]
        ed = end_date.split("T")[0]
        date_str = f"Periode: {sd} s/d {ed}"
    elif start_date:
        sd = start_date.split("T")[0]
        date_str = f"Periode: Mulai {sd}"
    elif end_date:
        ed = end_date.split("T")[0]
        date_str = f"Periode: Sampai {ed}"
        
    ws['A2'] = date_str
    ws['A2'].font = meta_font
    ws.merge_cells('A2:E2')
    
    ws['A3'] = f"Unduh Pada: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    ws['A3'].font = meta_font
    ws.merge_cells('A3:E3')
    
    ws.append([]) # Spacer

    headers = ["No. Faktur", "Tanggal & Waktu", "Supplier", "Status Pembayaran", "Total Belanja"]
    ws.append(headers)
    header_row = ws.max_row
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
    ws.row_dimensions[header_row].height = 25

    total_spending = 0.0

    for p in purchases:
        supp_name = p.supplier.name if p.supplier else "Supplier Umum"
        dt_local = p.purchase_date.strftime("%d-%m-%Y %H:%M:%S") if p.purchase_date else ""
        row_data = [
            f"BELI-{p.id}",
            dt_local,
            supp_name,
            p.payment_status,
            p.total_amount
        ]
        ws.append(row_data)
        curr_row = ws.max_row
        total_spending += p.total_amount
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
            if col_idx in [1, 2, 4]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="left")
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '"Rp "#,##0'

    ws.append([]) # Spacer row
    summary_row = ws.max_row + 1
    
    ws.cell(row=summary_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
    
    ws.cell(row=summary_row, column=5, value=total_spending).font = total_font
    ws.cell(row=summary_row, column=5).number_format = '"Rp "#,##0'
    ws.cell(row=summary_row, column=5).alignment = Alignment(horizontal="right")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.border = summary_border
        
    ws.row_dimensions[summary_row].height = 20

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 4 and cell.value:
                val_str = str(cell.value)
                if isinstance(cell.value, (int, float)):
                    val_str = f"Rp {cell.value:,.0f}"
                max_len = max(max_len, len(val_str))
            elif cell.row == 4 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    fn_date = datetime.datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="laporan_pembelian_{fn_date}.xlsx"'}
    )

@app.get("/api/reports/purchases/export/pdf")
def export_purchases_report_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.Purchase)
    if start_date:
        try:
            query = query.filter(models.Purchase.purchase_date >= datetime.datetime.fromisoformat(start_date.replace("Z", "+00:00")))
        except Exception:
            pass
    if end_date:
        try:
            query = query.filter(models.Purchase.purchase_date <= datetime.datetime.fromisoformat(end_date.replace("Z", "+00:00")))
        except Exception:
            pass

    purchases = query.order_by(models.Purchase.purchase_date.desc()).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=60
    )

    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#1A365D"),
        fontName="Helvetica-Bold",
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'ReportSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#4A5568"),
        fontName="Helvetica",
        spaceAfter=15
    )
    th_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=colors.white,
        fontName="Helvetica-Bold",
        alignment=1
    )
    td_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#2D3748"),
        fontName="Helvetica"
    )
    td_center_style = ParagraphStyle(
        'TableCellCenter',
        parent=td_style,
        alignment=1
    )
    td_right_style = ParagraphStyle(
        'TableCellRight',
        parent=td_style,
        alignment=2
    )

    story = []

    story.append(Paragraph("LAPORAN PEMBELIAN", title_style))
    
    date_str = "Semua Periode"
    if start_date and end_date:
        sd = start_date.split("T")[0]
        ed = end_date.split("T")[0]
        date_str = f"Periode: {sd} s/d {ed}"
    elif start_date:
        sd = start_date.split("T")[0]
        date_str = f"Periode: Mulai {sd}"
    elif end_date:
        ed = end_date.split("T")[0]
        date_str = f"Periode: Sampai {ed}"
        
    meta_text = f"<b>{date_str}</b> &nbsp;&nbsp;|&nbsp;&nbsp; Diunduh: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    story.append(Paragraph(meta_text, subtitle_style))

    table_data = [[
        Paragraph("No. Faktur", th_style),
        Paragraph("Tanggal & Waktu", th_style),
        Paragraph("Supplier", th_style),
        Paragraph("Status Pembayaran", th_style),
        Paragraph("Total Belanja", th_style)
    ]]

    total_spending = 0.0

    for p in purchases:
        supp_name = p.supplier.name if p.supplier else "Supplier Umum"
        dt_str = p.purchase_date.strftime("%d-%m-%Y %H:%M:%S") if p.purchase_date else ""
        table_data.append([
            Paragraph(f"BELI-{p.id}", td_center_style),
            Paragraph(dt_str, td_center_style),
            Paragraph(supp_name, td_style),
            Paragraph(p.payment_status or "", td_center_style),
            Paragraph(f"Rp {p.total_amount:,.0f}", td_right_style)
        ])
        total_spending += p.total_amount

    summary_idx = len(table_data)
    table_data.append([
        Paragraph("<b>TOTAL</b>", td_style),
        "", "", "",
        Paragraph(f"<b>Rp {total_spending:,.0f}</b>", td_right_style)
    ])

    t = Table(table_data, colWidths=[70, 110, 160, 90, 102])
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1A365D")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ])
    
    for i in range(1, summary_idx):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F7FAFC"))
            
    t_style.add('SPAN', (0, summary_idx), (3, summary_idx))
    t_style.add('BACKGROUND', (0, summary_idx), (-1, summary_idx), colors.HexColor("#EDF2F7"))
    t_style.add('LINEABOVE', (0, summary_idx), (-1, summary_idx), 1, colors.HexColor("#1A365D"))
    t_style.add('LINEBELOW', (0, summary_idx), (-1, summary_idx), 1.5, colors.HexColor("#1A365D"))

    t.setStyle(t_style)
    story.append(t)

    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    buffer.seek(0)
    
    fn_date = datetime.datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="laporan_pembelian_{fn_date}.pdf"'}
    )

@app.get("/api/reports/profit-loss/export/excel")
def export_profit_loss_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    pl_data = get_profit_loss_report(db, start_date, end_date)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Laba Rugi"
    ws.views.sheetView[0].showGridLines = True

    title_font = Font(name="Calibri", size=16, bold=True, color="1A365D")
    meta_font = Font(name="Calibri", size=10, italic=True, color="555555")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    double_bottom = Border(
        top=Side(style='thin', color='1A365D'),
        bottom=Side(style='double', color='1A365D')
    )

    ws['A1'] = "LAPORAN LABA RUGI (P&L)"
    ws['A1'].font = title_font
    ws.merge_cells('A1:C1')
    
    date_str = "Semua Periode"
    if start_date and end_date:
        sd = start_date.split("T")[0]
        ed = end_date.split("T")[0]
        date_str = f"Periode: {sd} s/d {ed}"
    elif start_date:
        sd = start_date.split("T")[0]
        date_str = f"Periode: Mulai {sd}"
    elif end_date:
        ed = end_date.split("T")[0]
        date_str = f"Periode: Sampai {ed}"
        
    ws['A2'] = date_str
    ws['A2'].font = meta_font
    ws.merge_cells('A2:C2')
    
    ws['A3'] = f"Unduh Pada: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    ws['A3'].font = meta_font
    ws.merge_cells('A3:C3')
    
    ws.append([]) # Spacer

    headers = ["Keterangan", "Nilai", "Subtotal"]
    ws.append(headers)
    header_row = ws.max_row
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
    ws.row_dimensions[header_row].height = 25

    ws.append(["Penjualan Bersih (Revenue)", "", pl_data["revenue"]])
    r1 = ws.max_row
    ws.cell(row=r1, column=1).font = bold_font
    ws.cell(row=r1, column=3).font = bold_font
    
    ws.append(["Harga Pokok Penjualan (HPP / COGS)", pl_data["cogs"], ""])
    r2 = ws.max_row
    ws.cell(row=r2, column=1).font = regular_font
    ws.cell(row=r2, column=2).font = regular_font
    
    ws.append(["LABA KOTOR (Gross Profit)", "", pl_data["gross_profit"]])
    r3 = ws.max_row
    ws.cell(row=r3, column=1).font = bold_font
    ws.cell(row=r3, column=3).font = bold_font
    ws.cell(row=r3, column=1).border = Border(top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
    ws.cell(row=r3, column=2).border = Border(top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
    ws.cell(row=r3, column=3).border = Border(top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
    
    ws.append(["Beban Operasional & Pengeluaran Kas", pl_data["expenses"], ""])
    r4 = ws.max_row
    ws.cell(row=r4, column=1).font = regular_font
    ws.cell(row=r4, column=2).font = regular_font
    
    ws.append(["LABA BERSIH (Net Profit)", "", pl_data["net_profit"]])
    r5 = ws.max_row
    ws.cell(row=r5, column=1).font = Font(name="Calibri", size=11, bold=True, color="1A365D" if pl_data["net_profit"] >= 0 else "9B2C2C")
    ws.cell(row=r5, column=3).font = Font(name="Calibri", size=11, bold=True, color="1A365D" if pl_data["net_profit"] >= 0 else "9B2C2C")
    ws.cell(row=r5, column=1).border = double_bottom
    ws.cell(row=r5, column=2).border = double_bottom
    ws.cell(row=r5, column=3).border = double_bottom

    for row_idx in [r1, r2, r3, r4, r5]:
        ws.row_dimensions[row_idx].height = 20
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right")
        
        if ws.cell(row=row_idx, column=2).value != "":
            ws.cell(row=row_idx, column=2).number_format = '"Rp "#,##0'
        if ws.cell(row=row_idx, column=3).value != "":
            ws.cell(row=row_idx, column=3).number_format = '"Rp "#,##0'
            
        for col_idx in [1, 2, 3]:
            cell = ws.cell(row=row_idx, column=col_idx)
            if not cell.border.top:
                cell.border = thin_border

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    fn_date = datetime.datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="laporan_laba_rugi_{fn_date}.xlsx"'}
    )

@app.get("/api/reports/profit-loss/export/pdf")
def export_profit_loss_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    pl_data = get_profit_loss_report(db, start_date, end_date)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=60
    )

    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#1A365D"),
        fontName="Helvetica-Bold",
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'ReportSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#4A5568"),
        fontName="Helvetica",
        spaceAfter=25
    )
    th_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.white,
        fontName="Helvetica-Bold",
        alignment=1
    )
    td_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#2D3748"),
        fontName="Helvetica"
    )
    td_bold_style = ParagraphStyle(
        'TableCellBold',
        parent=td_style,
        fontName="Helvetica-Bold"
    )
    td_right_style = ParagraphStyle(
        'TableCellRight',
        parent=td_style,
        alignment=2
    )
    td_right_bold_style = ParagraphStyle(
        'TableCellRightBold',
        parent=td_right_style,
        fontName="Helvetica-Bold"
    )

    story = []

    story.append(Paragraph("LAPORAN LABA RUGI (P&L)", title_style))
    
    date_str = "Semua Periode"
    if start_date and end_date:
        sd = start_date.split("T")[0]
        ed = end_date.split("T")[0]
        date_str = f"Periode: {sd} s/d {ed}"
    elif start_date:
        sd = start_date.split("T")[0]
        date_str = f"Periode: Mulai {sd}"
    elif end_date:
        ed = end_date.split("T")[0]
        date_str = f"Periode: Sampai {ed}"
        
    meta_text = f"<b>{date_str}</b> &nbsp;&nbsp;|&nbsp;&nbsp; Diunduh: {datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    story.append(Paragraph(meta_text, subtitle_style))

    table_data = [
        [
            Paragraph("Keterangan", th_style),
            Paragraph("Nilai", th_style),
            Paragraph("Subtotal", th_style)
        ],
        [
            Paragraph("Penjualan Bersih (Revenue)", td_bold_style),
            "",
            Paragraph(f"Rp {pl_data['revenue']:,.0f}", td_right_bold_style)
        ],
        [
            Paragraph("Harga Pokok Penjualan (HPP / COGS)", td_style),
            Paragraph(f"-Rp {pl_data['cogs']:,.0f}", td_right_style),
            ""
        ],
        [
            Paragraph("<b>LABA KOTOR (Gross Profit)</b>", td_bold_style),
            "",
            Paragraph(f"Rp {pl_data['gross_profit']:,.0f}", td_right_bold_style)
        ],
        [
            Paragraph("Beban Operasional & Pengeluaran Kas", td_style),
            Paragraph(f"-Rp {pl_data['expenses']:,.0f}", td_right_style),
            ""
        ],
        [
            Paragraph("<b>LABA BERSIH (Net Profit)</b>", td_bold_style),
            "",
            Paragraph(f"Rp {pl_data['net_profit']:,.0f}", td_right_bold_style)
        ]
    ]

    t = Table(table_data, colWidths=[320, 106, 106])
    
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1A365D")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor("#EDF2F7")),
        ('LINEABOVE', (0, 3), (-1, 3), 1, colors.HexColor("#1A365D")),
        ('LINEBELOW', (0, 3), (-1, 3), 1, colors.HexColor("#1A365D")),
        
        ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor("#E2E8F0")),
        ('LINEABOVE', (0, 5), (-1, 5), 1, colors.HexColor("#1A365D")),
        ('LINEBELOW', (0, 5), (-1, 5), 2, colors.HexColor("#1A365D")),
    ])
    
    t.setStyle(t_style)
    story.append(t)

    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    buffer.seek(0)
    
    fn_date = datetime.datetime.now().strftime("%Y%m%d")
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="laporan_laba_rugi_{fn_date}.pdf"'}
    )


# Cleaned up alias endpoints that are now integrated natively.

